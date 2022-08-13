import openpyxl



# File --> Workbook --> Sheet -> Rows --> Cells

# Writting same data
# file = "/home/prashant/Documents/test_data.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook["Sheet1"]  # or workbok.active  -- get active sheet form excel
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r, c).value = "Welcome"
# workbook.save(file)
# print("Successfully Written to file")



# Writting different data
file = "/home/prashant/Documents/test_data1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]  # or workbok.active  -- get active sheet form excel

sheet.cell(1,1).value = "Name"
sheet.cell(1,2).value = "ID"
sheet.cell(2,2).value = 123
sheet.cell(3,2).value = 456
sheet.cell(4,2).value = 678
sheet.cell(2,1).value = "Prashant"
sheet.cell(3,1).value = "Sachin"
sheet.cell(4,1).value = "Parag"

workbook.save(file)
print("Successfully written different data")
