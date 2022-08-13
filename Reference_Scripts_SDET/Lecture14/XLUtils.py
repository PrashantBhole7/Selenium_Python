import openpyxl
from openpyxl.styles import PatternFill

def get_row_count(file, sheetname):
    workbook  = openpyxl.load_workbook(file)
    sheet  = workbook[sheetname]
    return sheet.max_row


def get_column_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def read_data(file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=colnum).value


def write_data(file, sheetname, rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(file)


def fill_green_colour(file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    green_feel = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum, colnum).fill = green_feel
    workbook.save(file)


def fill_red_colour(file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    red_feel = PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(rownum, colnum).fill = red_feel
    workbook.save(file)
